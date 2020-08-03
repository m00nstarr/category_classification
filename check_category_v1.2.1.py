#################################################################################
# coded by Mun Hyung Lee / Email : moonstar114@naver.com , Phone : 010-4001-9614

# Industry field definition posted on gdrive/2020intern/
# code last update : 2020 / 08/03

# How to use
# > python3.5 check_category.py
# > (File_name).xlsx   (insert file_name)


import pandas as pd
import numpy as np
import openpyxl
import time
import math
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string
import os

print("File name >")
file_name = input()
path = os.path.normpath(os.getcwd())


KR_2020 = pd.read_excel( path +"/"+file_name+ ".xlsx", sheet_name= 'RAW')
print()
print("** Read "+ str(KR_2020.shape[0]) + " records of "+ file_name+ ".xlsx")  

workbook = openpyxl.load_workbook(path +"/" +file_name+".xlsx")
worksheet = workbook['RAW']


print("** generating outputfile ....")
create_time = time.strftime('%H%M', time.localtime(time.time()))
new_workbook = Workbook(file_name+"_out_"+ create_time +".xlsx")

print("** "+ file_name +"_out_"+create_time+".xlsx"+" generated")
print()
new_worksheet = new_workbook.create_sheet('RAW')
new_workbook.close()
new_workbook.save(file_name+"_out_"+create_time+".xlsx")

new_workbook= openpyxl.load_workbook(path+ "/" +file_name+"_out_"+create_time+".xlsx")
new_worksheet = new_workbook['RAW']

contents = KR_2020[["subcategory","App","OS"]]
# extracts only these three cols

need_edit_set = {}
need_edit_set = set()
null_edit_set = {}
null_edit_set = set()
null_index_list = []  # update 20/07/21 : for notifying where is null row

for row in worksheet:
  for cell in row:
    new_worksheet[cell.coordinate].value = cell.value

print("** copying original to outputfile completed")

null_count=0
for i in contents[contents['subcategory'].isnull()].index:
  null_edit_set.add(contents['App'][i])
  null_count = null_count +1 
  null_index_list.append(i)


null_edit = []
null_edit = list(null_edit_set)

# write maincategory using subcategory
count =0
for i in range(len(contents)):
  sub_name = contents['subcategory'][i]
  os_name = contents['OS'][i]

  if sub_name == "Sports":
    need_edit_set.add( contents['App'][i])
    new_worksheet.cell(row=i + 2, column=column_index_from_string('i')).value = "Sports"

  elif sub_name == "Action" or sub_name =="Adventure" \
    or sub_name =="Arcade" or sub_name == "Board" \
    or sub_name =="Card" or sub_name == "Casino" or sub_name == "Casual"\
    or sub_name == "Puzzle" or sub_name == "Racing" or sub_name == "Role Playing"\
    or sub_name == "Simulation" or sub_name == "Sportsgame" \
    or sub_name =="Strategy" or sub_name =="Trivia" or sub_name == "Word" or (sub_name =="Music" and os_name == "AND") :
    
    new_worksheet.cell(row=i + 2, column=column_index_from_string('i')).value = "Games"

  elif sub_name == "Travel & Local": 
    new_worksheet.cell(row=i + 2, column=column_index_from_string('i')).value = "Travel"

  elif sub_name == "Books" or sub_name == "Reference":
    new_worksheet.cell(row=i + 2, column=column_index_from_string('i')).value = "Books & Reference"

  elif sub_name == "Events" or sub_name == "Family" or sub_name =="House & Home" or sub_name == "Parenting" \
    or sub_name == "Auto & Vehicles" or sub_name == "Beauty" or sub_name == "Weather" :
    new_worksheet.cell(row=i + 2, column=column_index_from_string('i')).value = "Lifestyle"

  elif sub_name == "Health & Fitness" or sub_name == "Healthcare & Fitness" or sub_name == "Medical":
    new_worksheet.cell(row=i + 2, column=column_index_from_string('i')).value = "Health"

  elif sub_name == "Communication" or sub_name =="Dating" or sub_name =="Social Networking" or sub_name =="Events":
    new_worksheet.cell(row=i + 2, column=column_index_from_string('i')).value = "Social"

  elif sub_name == "Photography" or sub_name == "Video Players & Editors":
    new_worksheet.cell(row=i + 2, column=column_index_from_string('i')).value = "Photo & Video"

  elif sub_name == "News & Magazines" or sub_name == "Magazines & Newspapers":
    new_worksheet.cell(row=i + 2, column=column_index_from_string('i')).value = "News"

  elif sub_name == "Tools" or sub_name =="Libraries & Demo" or sub_name =="Productivity" or sub_name =="Art & Design" or sub_name == "Personalization":
    new_worksheet.cell(row=i + 2, column=column_index_from_string('i')).value = "Utilities"

  elif sub_name == "Maps & Navigations":
    new_worksheet.cell(row=i + 2, column=column_index_from_string('i')).value = "Navigation"

  elif sub_name == "Music & Audio" or sub_name =="Music":
    new_worksheet.cell(row=i + 2, column=column_index_from_string('i')).value = "Music"
    if os_name == "IOS":
      need_edit_set.add(contents['App'][i])
      
  elif sub_name == "Comics":
    new_worksheet.cell(row=i + 2, column=column_index_from_string('i')).value = "Entertainment"
  elif sub_name == "Kids" or sub_name == "Educational":
    new_worksheet.cell(row=i + 2, column=column_index_from_string('i')).value = "Education"
    if sub_name == "Educational":
      need_edit_set.add(contents['App'][i])
  else:
    new_worksheet.cell(row=i + 2, column=column_index_from_string('i')).value = sub_name
  
  count = count +1 

print()
print("** fills industry filed of " +str(count)+ " records")


if count != KR_2020.shape[0]:
  blank_count = KR_2020.shape[0] - count +1 
  print("There are" + str(blank_count) + " blank(s) industry field")
  print("need to check")
# sports need to be edited. (mixed subcategory)


# make list of confused application ( sports category )
need_edit = []
need_edit = list(need_edit_set)



print()
print("updating confusing application")
print("==============================")

for app_name in need_edit:
  print()
  print("'"+app_name +"'")
  print("Game ? (y/n) " )

  while True:
    ans = input()

    if ans == "y":
      for i in np.where(KR_2020["App"] == app_name)[0]:
        new_worksheet.cell(row =i+2, column = column_index_from_string('i')).value = "Games"
      break

    elif ans == "n":
      break

    else:
      print("press y/n")



print() 
print("There are "+ str(null_count) + " null subcategory application records") 
print("---------------------------------------------------------------------")

for index in null_index_list:
  print("check " + str(index+2) + " th row of file  << Null row ")
# 20/07/21 null row detection

for app_name in null_edit:
  
  if type(app_name) is str:
    if not app_name:
      continue  #preventing from error / null detect
  else :
    if math.isnan(app_name):
      continue  #preventing from error / null detect

  print("* Write industry field of >> " + app_name)
  Maincategory = input()

  count =0
  for i in np.where(KR_2020["App"]==app_name)[0]:
    new_worksheet.cell(row =i+2, column = column_index_from_string('i')).value = Maincategory
    count = count + 1
  print()
  print( "fill industry field of >> " + str(count) +" " +app_name + " record(s)")#
  print()
 
  print("* Write Subcategory of " + app_name) 
  Subcategory = input()
  count =0
  for i in np.where(KR_2020["App"]== app_name)[0]:
    new_worksheet.cell(row = i+2, column = column_index_from_string('j')).value = Subcategory
    count = count +1
  
  print()
  print( "fill Subcategory field of "+ str(count) + " "+ app_name + " record(s)")#

new_workbook.close()
new_workbook.save(file_name+ "_out_"+create_time+".xlsx")

print()
print("Update complete")
